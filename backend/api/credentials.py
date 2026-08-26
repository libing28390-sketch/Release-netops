"""Credential vault API with two-phase hardware password synchronization."""

import html
import io
import logging
import zipfile

from fastapi import APIRouter, BackgroundTasks, Body, HTTPException, Query
from fastapi.responses import JSONResponse, Response

from core.rbac import require_role
from database import get_db_connection
from schemas.schemas import CredentialCreate, CredentialUpdate
from services import credential_service
from services.audit_service import log_audit_event

logger = logging.getLogger(__name__)

router = APIRouter()

CREDENTIAL_TEMPLATE_VERSION = '2'


def _build_credential_import_template() -> bytes:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = openpyxl.Workbook()

    # Sheet 1: SSH
    ws_ssh = wb.active
    ws_ssh.title = 'SSH'

    # Sheet 2: SNMP
    ws_snmp = wb.create_sheet(title='SNMP')

    header_fill = PatternFill(start_color='FF007D9D', end_color='FF007D9D', fill_type='solid')
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFFFF')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    header_border = Border(
        left=Side(style='thin', color='FFB7D7E0'),
        right=Side(style='thin', color='FFB7D7E0'),
        top=Side(style='thin', color='FFB7D7E0'),
        bottom=Side(style='thin', color='FFB7D7E0'),
    )

    ssh_columns = [
        ('凭据名称', 24), ('类型', 18), ('用户名', 18), ('账号角色', 18),
        ('密码', 24), ('Enable 密码', 24),
    ]
    snmp_columns = [
        ('凭据名称', 24), ('类型', 18), ('SNMP 团体字', 26), ('SNMP 服务器地址', 28),
    ]

    # Configure SSH sheet
    ws_ssh.append([label for label, _width in ssh_columns])
    ws_ssh.row_dimensions[1].height = 24
    ws_ssh.freeze_panes = 'A2'
    ws_ssh.auto_filter.ref = 'A1:F1000'
    for col_idx, (_label, width) in enumerate(ssh_columns, start=1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        ws_ssh.column_dimensions[col_letter].width = width
        cell = ws_ssh.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = header_border

    dv_ssh_type = DataValidation(type='list', formula1='"ssh_password,ssh_key,api_token"', allow_blank=True)
    dv_ssh_type.errorTitle = '输入值不合法'
    dv_ssh_type.error = '请选择下拉选项'
    dv_ssh_type.promptTitle = '可选值'
    dv_ssh_type.prompt = '请从下拉列表选择'
    ws_ssh.add_data_validation(dv_ssh_type)
    dv_ssh_type.add('B2:B1000')

    dv_ssh_role = DataValidation(type='list', formula1='"普通账号,特权账号,通用账号,未指定"', allow_blank=True)
    dv_ssh_role.errorTitle = '输入值不合法'
    dv_ssh_role.error = '请选择下拉选项'
    dv_ssh_role.promptTitle = '可选值'
    dv_ssh_role.prompt = '请从下拉列表选择'
    ws_ssh.add_data_validation(dv_ssh_role)
    dv_ssh_role.add('D2:D1000')

    # Configure SNMP sheet
    ws_snmp.append([label for label, _width in snmp_columns])
    ws_snmp.row_dimensions[1].height = 24
    ws_snmp.freeze_panes = 'A2'
    ws_snmp.auto_filter.ref = 'A1:D1000'
    for col_idx, (_label, width) in enumerate(snmp_columns, start=1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        ws_snmp.column_dimensions[col_letter].width = width
        cell = ws_snmp.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = header_border

    dv_snmp_type = DataValidation(type='list', formula1='"snmpv2,snmpv3"', allow_blank=True)
    dv_snmp_type.errorTitle = '输入值不合法'
    dv_snmp_type.error = '请选择下拉选项'
    dv_snmp_type.promptTitle = '可选值'
    dv_snmp_type.prompt = '请从下拉列表选择'
    ws_snmp.add_data_validation(dv_snmp_type)
    dv_snmp_type.add('B2:B1000')

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


@router.get('/credentials/import-template', response_class=Response)
def api_download_credential_import_template(_user=require_role('Operator')):
    return Response(
        content=_build_credential_import_template(),
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={
            'Content-Disposition': 'attachment; filename="credential_import_template.xlsx"',
            'Cache-Control': 'no-store',
            'X-Credential-Template-Version': CREDENTIAL_TEMPLATE_VERSION,
        },
    )


_REVEALABLE_SECRET_COLUMNS = {
    "password": "encrypted_password",
    "enable_password": "enable_password",
}


def _credential_secret_column(secret_type: str) -> str:
    column = _REVEALABLE_SECRET_COLUMNS.get((secret_type or "").strip().lower())
    if not column:
        raise HTTPException(status_code=400, detail="Unsupported credential secret type")
    return column


def _password_change_has_binding_changes(conn, cred_id: str, submitted: dict) -> bool:
    """Return True only when username/account role actually changed."""
    current = credential_service.get_credential(conn, cred_id)
    for field in ("username", "account_role"):
        if field not in submitted:
            continue
        current_value = str(current.get(field) or '').strip()
        incoming = str(submitted[field] or '').strip()
        if field == "account_role":
            current_value = current_value.lower()
            incoming = incoming.lower()
        if current_value != incoming:
            return True
    return False


@router.get("/credentials")
def api_list_credentials(_user=require_role("Operator")):
    conn = get_db_connection()
    try:
        data = credential_service.list_credentials(conn)
        return {"success": True, "data": data, "message": ""}
    finally:
        conn.close()


@router.get("/credentials/{cred_id}")
def api_get_credential(cred_id: str, _user=require_role("Operator")):
    conn = get_db_connection()
    try:
        cred = credential_service.get_credential(conn, cred_id)
        return {"success": True, "data": cred, "message": ""}
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    finally:
        conn.close()


@router.get("/credentials/{cred_id}/secret")
def api_reveal_credential_secret(
    cred_id: str,
    secret_type: str = Query(default="password", alias="type"),
    user=require_role("Administrator"),
):
    """Reveal the authoritative vault secret for administrators only."""
    from core.crypto import decrypt_credential

    column = _credential_secret_column(secret_type)
    conn = get_db_connection()
    try:
        row = conn.execute(
            f"SELECT credential_name, {column} AS encrypted_secret FROM credentials WHERE id = ?",
            (cred_id,),
        ).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Credential not found")
        encrypted_secret = row["encrypted_secret"] or ""
        if not encrypted_secret:
            raise HTTPException(status_code=404, detail="Credential secret is not configured")
        secret = decrypt_credential(encrypted_secret)
        if secret is None:
            raise HTTPException(status_code=500, detail="Credential secret could not be decrypted")

        log_audit_event(
            event_type="credential.secret.view",
            category="security",
            severity="high",
            status="success",
            summary=f"Viewed {secret_type} for credential {cred_id}",
            actor_username=user.get("username") if isinstance(user, dict) else None,
            actor_role=user.get("role") if isinstance(user, dict) else None,
            target_type="credential",
            target_id=cred_id,
            target_name=row["credential_name"],
            details={"secret_type": secret_type, "action": "view"},
            conn=conn,
        )
        conn.commit()
        return {"success": True, "data": {"secret": secret, "secret_type": secret_type}}
    finally:
        conn.close()


@router.post("/credentials/{cred_id}/secret/copy")
def api_audit_credential_secret_copy(
    cred_id: str,
    secret_type: str = Query(default="password", alias="type"),
    user=require_role("Administrator"),
):
    """Record an administrator copying a previously revealed vault secret."""
    column = _credential_secret_column(secret_type)
    conn = get_db_connection()
    try:
        row = conn.execute(
            f"SELECT credential_name, {column} AS encrypted_secret FROM credentials WHERE id = ?",
            (cred_id,),
        ).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Credential not found")
        if not row["encrypted_secret"]:
            raise HTTPException(status_code=404, detail="Credential secret is not configured")

        log_audit_event(
            event_type="credential.secret.copy",
            category="security",
            severity="high",
            status="success",
            summary=f"Copied {secret_type} for credential {cred_id}",
            actor_username=user.get("username") if isinstance(user, dict) else None,
            actor_role=user.get("role") if isinstance(user, dict) else None,
            target_type="credential",
            target_id=cred_id,
            target_name=row["credential_name"],
            details={"secret_type": secret_type, "action": "copy"},
            conn=conn,
        )
        conn.commit()
        return {"success": True, "data": None, "message": "Secret copy audited"}
    finally:
        conn.close()


@router.post("/credentials/{cred_id}/bindings/{device_id}")
def api_manage_credential_binding(
    cred_id: str,
    device_id: str,
    body: dict = Body(...),
    user=require_role("Administrator"),
):
    """Replace or clear one device binding without deleting the credential."""
    action = str(body.get("action") or "").strip().lower()
    if action not in {"replace", "unbind"}:
        raise HTTPException(status_code=400, detail="Action must be replace or unbind")

    conn = get_db_connection()
    try:
        device = conn.execute(
            "SELECT id, asset_id, hostname, credential_id, admin_credential_id FROM devices WHERE id = ?",
            (device_id,),
        ).fetchone()
        if not device:
            raise HTTPException(status_code=404, detail="Device not found")
        device = dict(device)

        if str(device.get("credential_id") or "") == cred_id:
            binding_column = "credential_id"
            asset_column = "credential_id"
            binding_role = "normal"
        elif str(device.get("admin_credential_id") or "") == cred_id:
            binding_column = "admin_credential_id"
            asset_column = "admin_credential_id"
            binding_role = "admin"
        else:
            raise HTTPException(status_code=409, detail="Credential is not bound to this device")

        # Do not detach a target while a credential-wide synchronization job
        # is changing hardware. The lock is created by the durable sync flow.
        try:
            lock = conn.execute(
                "SELECT job_id FROM credential_password_sync_locks WHERE credential_id = ?",
                (cred_id,),
            ).fetchone()
        except Exception:
            lock = None
        if lock:
            raise HTTPException(status_code=409, detail="Credential synchronization is running; wait for it to finish before changing bindings")

        replacement_id = ""
        if action == "replace":
            replacement_id = str(body.get("replacement_credential_id") or "").strip()
            if not replacement_id or replacement_id == cred_id:
                raise HTTPException(status_code=400, detail="A different replacement credential is required")
            replacement = credential_service.get_credential(conn, replacement_id)
            replacement_role = str(replacement.get("account_role") or "").lower()
            if binding_role == "admin" and replacement_role == "normal":
                raise HTTPException(status_code=422, detail="A privileged device binding requires a privileged-compatible credential")
            if binding_role == "normal" and replacement_role == "admin":
                raise HTTPException(status_code=422, detail="A normal device binding cannot use a privileged-only credential")

        new_value = replacement_id if action == "replace" else None
        conn.execute(f"UPDATE devices SET {binding_column} = ? WHERE id = ?", (new_value, device_id))
        if device.get("asset_id"):
            conn.execute(
                f"UPDATE physical_assets SET {asset_column} = ? WHERE id = ?",
                (replacement_id if action == "replace" else "", device["asset_id"]),
            )
        conn.commit()

        action_label = "Replaced" if action == "replace" else "Unbound"
        log_audit_event(
            event_type="credential.binding.replace" if action == "replace" else "credential.binding.unbind",
            category="security",
            severity="high",
            status="success",
            summary=f"{action_label} {binding_role} credential binding for device {device_id}",
            actor_username=user.get("username") if isinstance(user, dict) else None,
            actor_role=user.get("role") if isinstance(user, dict) else None,
            target_type="device",
            target_id=device_id,
            target_name=device.get("hostname"),
            device_id=device_id,
            details={
                "binding_role": binding_role,
                "source_credential_id": cred_id,
                "replacement_credential_id": replacement_id if action == "replace" else None,
                "mode": "independent_device_credential" if action == "unbind" else "replacement_credential",
            },
        )
        return {
            "success": True,
            "data": {
                "device_id": device_id,
                "action": action,
                "binding_role": binding_role,
                "replacement_credential_id": replacement_id or None,
            },
            "message": "Credential binding updated",
        }
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    finally:
        conn.close()


@router.post("/credentials", status_code=201)
def api_create_credential(body: CredentialCreate, user=require_role("Operator")):
    conn = get_db_connection()
    try:
        cred = credential_service.create_credential(conn, **body.model_dump())
        log_audit_event(
            event_type="credential.create",
            category="configuration",
            severity="info",
            status="success",
            summary=f"Created credential '{body.credential_name}'",
            actor_username=user.get("username") if isinstance(user, dict) else None,
            actor_role=user.get("role") if isinstance(user, dict) else None,
            target_type="credential",
            target_id=cred.get("id"),
            target_name=body.credential_name,
        )
        return {"success": True, "data": cred, "message": "Credential created"}
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    finally:
        conn.close()


@router.put("/credentials/{cred_id}")
def api_update_credential(
    cred_id: str,
    body: CredentialUpdate,
    background_tasks: BackgroundTasks,
    user=require_role("Operator"),
):
    conn = get_db_connection()
    try:
        submitted = body.model_dump(exclude_unset=True)
        new_password = submitted.pop("password", None)
        new_enable_password = submitted.pop("enable_password", None)
        old_password = submitted.pop("old_password", None)
        old_enable_password = submitted.pop("old_enable_password", None)
        actor_username = user.get("username") if isinstance(user, dict) else None
        has_secret_change = new_password is not None or new_enable_password is not None

        if has_secret_change and _password_change_has_binding_changes(conn, cred_id, submitted):
            raise ValueError("Username or account role changes must be submitted separately from a password change")

        if has_secret_change:
            from services.credential_password_sync_service import (
                create_password_sync_job,
                run_credential_password_sync_job,
            )

            sync = create_password_sync_job(
                conn,
                credential_id=cred_id,
                new_password=new_password,
                new_enable_password=new_enable_password,
                old_password=old_password,
                old_enable_password=old_enable_password,
                actor_username=actor_username,
            )
            device_count = int(sync.get("device_count") or 0)
            if device_count > 0:
                # Metadata changes are independent of the secret transaction;
                # username/account-role changes were rejected above because
                # they would invalidate the hardware target snapshot.
                if submitted:
                    credential_service.update_credential(conn, cred_id, **submitted)
                cred = credential_service.get_credential(conn, cred_id)
                background_tasks.add_task(run_credential_password_sync_job, sync["job_id"])
                log_audit_event(
                    event_type="credential.password_sync.queued",
                    category="configuration",
                    severity="warning",
                    status="queued",
                    summary=f"Queued hardware password synchronization for credential {cred_id} on {device_count} devices",
                    actor_username=actor_username,
                    actor_role=user.get("role") if isinstance(user, dict) else None,
                    target_type="credential",
                    target_id=cred_id,
                    target_name=cred.get("credential_name"),
                    job_id=sync["job_id"],
                )
                return JSONResponse(
                    status_code=202,
                    content={
                        "success": True,
                        "data": cred,
                        "job_id": sync["job_id"],
                        "device_count": device_count,
                        "message": "Credential change accepted; hardware synchronization is running",
                    },
                )

            # No bound devices: validation succeeded, so the vault can be
            # updated directly without a hardware task.
            submitted["password"] = new_password
            submitted["enable_password"] = new_enable_password

        cred = credential_service.update_credential(conn, cred_id, **submitted)
        log_audit_event(
            event_type="credential.update",
            category="configuration",
            severity="info",
            status="success",
            summary=f"Updated credential {cred_id}",
            actor_username=actor_username,
            actor_role=user.get("role") if isinstance(user, dict) else None,
            target_type="credential",
            target_id=cred_id,
            target_name=cred.get("credential_name"),
        )
        return {"success": True, "data": cred, "message": "Credential updated"}
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    finally:
        conn.close()


@router.delete("/credentials/{cred_id}")
def api_delete_credential(cred_id: str, user=require_role("Administrator")):
    conn = get_db_connection()
    try:
        credential_service.delete_credential(conn, cred_id)
        log_audit_event(
            event_type="credential.delete",
            category="configuration",
            severity="warning",
            status="success",
            summary=f"Deleted credential {cred_id}",
            actor_username=user.get("username") if isinstance(user, dict) else None,
            actor_role=user.get("role") if isinstance(user, dict) else None,
            target_type="credential",
            target_id=cred_id,
        )
        return {"success": True, "data": None, "message": "Credential deleted"}
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    finally:
        conn.close()
