"""Build RackVision's reviewable asset catalog from the supplied workbook.

This is an offline import utility, not an application-startup dependency.  It
keeps the workbook unchanged and emits small, reviewable YAML/JSON files under
``assets/catalog``.  Missing dimensions remain ``null`` with an explicit
``pending_verification`` status; no geometry is inferred from images or from a
generic 1U default.

Example::

    python backend/scripts/build_rackvision_catalog.py \
        --source D:/data/downloads/Nexora_3D_Asset_Base_Catalog_v1.xlsx
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable


VENDOR_MAP = {
    "huawei": "huawei",
    "华为": "huawei",
    "h3c": "h3c",
    "新华三": "h3c",
    "ruijie": "ruijie",
    "锐捷": "ruijie",
    "zte": "zte",
    "中兴": "zte",
    "maipu": "maipu",
    "迈普": "maipu",
    "dptech": "dptech",
    "迪普": "dptech",
}

PORT_TOKEN_RE = re.compile(
    r"(?P<count>\d+)\s*[×x*]\s*(?P<spec>[^+，,;；]+)",
    re.IGNORECASE,
)


def clean(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return "" if text.lower() in {"nan", "none", "null", "tbd", "待核", "待确认"} else text


def nullable(value: Any) -> str | None:
    text = clean(value)
    return text or None


def number(value: Any) -> int | float | None:
    text = clean(value).replace(",", "")
    if not text:
        return None
    try:
        parsed = float(text)
    except ValueError:
        return None
    return int(parsed) if parsed.is_integer() else parsed


def parse_height(value: Any) -> int | None:
    parsed = number(value)
    if parsed is not None:
        return int(parsed)
    match = re.search(r"(?<!\d)(\d{1,2})\s*U", clean(value), re.IGNORECASE)
    return int(match.group(1)) if match else None


def parse_dimensions(value: Any) -> dict[str, int | float] | None:
    """Parse only explicit W x D x H text, never approximate wording."""

    text = clean(value)
    if not text:
        return None
    match = re.search(
        r"(?<!\d)(\d+(?:\.\d+)?)\s*[x×*]\s*"
        r"(\d+(?:\.\d+)?)\s*[x×*]\s*"
        r"(\d+(?:\.\d+)?)\s*(?:mm)?(?!\d)",
        text,
        re.IGNORECASE,
    )
    if not match:
        return None
    values = [number(item) for item in match.groups()]
    if any(item is None for item in values):
        return None
    return {"width_mm": values[0], "depth_mm": values[1], "height_mm": values[2]}


def parse_port_metadata(value: Any) -> dict[str, Any]:
    """Extract only explicit port counts; leave non-counted wording unknown."""

    text = clean(value)
    tokens = []
    for match in PORT_TOKEN_RE.finditer(text):
        count = int(match.group("count"))
        spec = match.group("spec").strip()
        tokens.append({"count": count, "spec": spec})
    if not tokens:
        return {
            "port_count": None,
            "port_media_type": "unknown",
            "port_breakdown": [],
            "uplink_ports": None,
            "management_ports": None,
            "console_ports": None,
        }

    lowered = " ".join(item["spec"].casefold() for item in tokens)
    has_copper = any(marker in lowered for marker in ("base-t", "rj45", "poe", "铜"))
    has_optical = any(marker in lowered for marker in ("sfp", "qsfp", "osfp", "光"))
    if has_copper and has_optical:
        media_type = "mixed"
    elif has_copper:
        media_type = "copper"
    elif has_optical:
        media_type = "optical"
    else:
        media_type = "ethernet"
    return {
        "port_count": sum(item["count"] for item in tokens),
        "port_media_type": media_type,
        "port_breakdown": tokens,
        "uplink_ports": None,
        "management_ports": None,
        "console_ports": None,
    }


def vendor(value: Any) -> str:
    raw = clean(value)
    return VENDOR_MAP.get(raw.casefold(), raw.casefold() or "generic")


def slug(value: Any) -> str:
    text = clean(value).casefold()
    parts = re.findall(r"[a-z0-9]+", text)
    return "_".join(parts) or "unknown"


def device_type(domain: str, role: str) -> str:
    text = f"{domain} {role}".casefold()
    rules = (
        (("防火墙", "firewall"), "firewall"),
        (("路由", "router"), "router"),
        (("服务器", "server"), "server"),
        (("存储", "storage"), "storage"),
        (("pdu", "pdu"), "pdu"),
        (("配线", "patch_panel"), "patch_panel"),
        (("交换", "switch"), "switch"),
        (("switch", "router", "firewall"), "switch"),
    )
    for markers, result in rules:
        if any(marker.casefold() in text for marker in markers):
            return result
    return "switch"


def mount_kind(install_location: str, asset_name: str) -> str:
    text = f"{install_location} {asset_name}".casefold()
    if "落地" in text or "floor" in text:
        return "floor"
    if "0u" in text or "pdu" in text:
        return "zero_u"
    if "机柜" in text or re.search(r"\d{2}U", text, re.IGNORECASE):
        return "u_mount"
    return "unknown"


def row_values(row: Iterable[Any], width: int) -> list[Any]:
    values = list(row)
    return values[:width] + [None] * max(0, width - len(values))


def source_meta(source_version: str, source_url: str, title: str) -> dict[str, Any]:
    return {
        "source_title": title or "Nexora 3D Asset Base Catalog",
        "source_url": source_url or None,
        "source_version": source_version or "unknown",
    }


def build_network(rows: list[tuple[int, list[Any]]]) -> tuple[list[dict[str, Any]], int]:
    records: list[dict[str, Any]] = []
    seen: dict[str, int] = {}
    incomplete = 0
    for source_row, values in rows:
        values = row_values(values, 17)
        maker = vendor(values[0])
        domain = nullable(values[1]) or "unknown"
        role = nullable(values[2]) or "unknown"
        family = nullable(values[3]) or "unknown"
        exact_model = nullable(values[4]) or f"row-{source_row}"
        dimensions = parse_dimensions(values[7])
        port_metadata = parse_port_metadata(values[5])
        if not dimensions:
            incomplete += 1
        base_key = f"{maker}__{slug(exact_model)}"
        seen[base_key] = seen.get(base_key, 0) + 1
        catalog_key = base_key if seen[base_key] == 1 else f"{base_key}__{seen[base_key]}"
        source_url = nullable(values[10]) or nullable(values[11])
        source_version = nullable(values[16]) or "unknown"
        record: dict[str, Any] = {
            "catalog_key": catalog_key,
            "vendor": maker,
            "family": family,
            "exact_model": exact_model,
            "device_type": device_type(domain, role),
            "deployment_domain": domain,
            "role": role,
            "height_u": parse_height(values[6]),
            "width_mm": dimensions["width_mm"] if dimensions else None,
            "depth_mm": dimensions["depth_mm"] if dimensions else None,
            "height_mm": dimensions["height_mm"] if dimensions else None,
            "dimension_status": "estimated" if dimensions else "pending_verification",
            "data_confidence": "verified_official" if source_url else "tbd",
            "port_layout": nullable(values[5]),
            **port_metadata,
            "psu_count": None,
            "fan_count": None,
            "blender_features": nullable(values[9]),
            "modeling_approach": nullable(values[8]) or "family_template",
            "asset_priority": nullable(values[13]) or "P2",
            "recommended_glb": nullable(values[14]),
            "notes": nullable(values[15]),
            "source_row": source_row,
            **source_meta(source_version, source_url or "", f"{maker} {family} official product page"),
        }
        records.append(record)
    return records, incomplete


def build_rack_assets(rows: list[tuple[int, list[Any]]]) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    seen: dict[str, int] = {}
    for source_row, values in rows:
        values = row_values(values, 10)
        asset_name = nullable(values[1]) or f"rack-asset-{source_row}"
        base_key = f"rack__{slug(asset_name)}"
        seen[base_key] = seen.get(base_key, 0) + 1
        key = base_key if seen[base_key] == 1 else f"{base_key}__{seen[base_key]}"
        records.append({
            "catalog_key": key,
            "category": nullable(values[0]) or "unknown",
            "asset_model": asset_name,
            "u_or_size": nullable(values[2]) or "unknown",
            "dimensions": nullable(values[3]) or "unknown",
            "mount_kind": mount_kind(clean(values[4]), asset_name),
            "installation_location": nullable(values[4]) or "unknown",
            "structural_features": nullable(values[5]),
            "nexora_fields": nullable(values[6]),
            "blender_approach": nullable(values[7]) or "generic_template",
            "asset_priority": nullable(values[8]) or "P2",
            "data_confidence": "verified_manual" if nullable(values[9]) else "tbd",
            "dimension_status": "pending_verification",
            "notes_or_source": nullable(values[9]),
            "source_row": source_row,
            **source_meta("unknown", "", "Nexora rack and accessory asset catalog"),
        })
    return records


def build_cables(rows: list[tuple[int, list[Any]]]) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    seen: dict[str, int] = {}
    for source_row, values in rows:
        values = row_values(values, 10)
        name = nullable(values[1]) or f"cable-optic-{source_row}"
        base_key = f"cable__{slug(name)}"
        seen[base_key] = seen.get(base_key, 0) + 1
        key = base_key if seen[base_key] == 1 else f"{base_key}__{seen[base_key]}"
        records.append({
            "catalog_key": key,
            "category": nullable(values[0]) or "unknown",
            "model": name,
            "connector_a": nullable(values[2]) or "unknown",
            "connector_b": nullable(values[3]) or "unknown",
            "supported_rate": nullable(values[4]) or "unknown",
            "length_options": nullable(values[5]) or "unknown",
            "blender_approach": nullable(values[6]) or "procedural_bezier",
            "nexora_data_mapping": nullable(values[7]) or "from_port/to_port",
            "market_note": nullable(values[8]),
            "asset_priority": nullable(values[9]) or "P2",
            "source_row": source_row,
            **source_meta("unknown", "", "Nexora cable and optic asset catalog"),
        })
    return records


def build_priorities(rows: list[tuple[int, list[Any]]]) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    for source_row, values in rows:
        values = row_values(values, 6)
        records.append({
            "stage": nullable(values[0]) or "unknown",
            "asset_or_model": nullable(values[1]) or f"row-{source_row}",
            "category": nullable(values[2]) or "unknown",
            "asset_hierarchy": nullable(values[3]) or "unknown",
            "rationale": nullable(values[4]),
            "acceptance_criteria": nullable(values[5]),
            "source_row": source_row,
        })
    return records


def build_registry(network: list[dict[str, Any]], rack_assets: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Create reusable family entries instead of one GLB per exact SKU."""

    grouped: dict[str, dict[str, Any]] = {}
    for record in network:
        glb_name = clean(record.get("recommended_glb"))
        asset_key = slug(Path(glb_name).stem) if glb_name else f"{record['vendor']}__{slug(record['family'])}__generic"
        entry = grouped.setdefault(asset_key, {
            "asset_key": asset_key,
            "vendor": record["vendor"],
            "family": record["family"],
            "exact_models": [],
            "device_type": record["device_type"],
            "deployment_domain": record["deployment_domain"],
            "height_u": record["height_u"],
            "width_mm": record["width_mm"],
            "depth_mm": record["depth_mm"],
            "height_mm": record["height_mm"],
            "dimension_status": record.get("dimension_status") or "pending_verification",
            "front_orientation": "front",
            "lod": ["LOD0", "LOD1", "LOD2"],
            "glb_path": f"assets/build/glb/{glb_name}" if glb_name else None,
            "public_path": f"/assets/3d/{glb_name}" if glb_name else None,
            "thumbnail_path": None,
            "source_blend": f"assets/source/blender/{asset_key}.blend",
            "source_refs": [],
            "data_confidence": record["data_confidence"],
            "version": "1.0.0",
            "status": "draft",
            "render_strategy": "procedural_until_glb_approved",
        })
        if record["exact_model"] not in entry["exact_models"]:
            entry["exact_models"].append(record["exact_model"])
        if record.get("source_url") and record["source_url"] not in entry["source_refs"]:
            entry["source_refs"].append(record["source_url"])
        # A shared family asset is only dimension-confirmed when every grouped
        # SKU has explicit dimensions; otherwise the resolver remains honest.
        for field in ("width_mm", "depth_mm", "height_mm"):
            if entry[field] != record[field]:
                entry[field] = None
                entry["dimension_status"] = "pending_verification"
        if entry["data_confidence"] != record["data_confidence"]:
            entry["data_confidence"] = "tbd"

    generic_templates = [
        ("generic_switch_1u_24p", "generic", "switch_1u", "switch", 1),
        ("generic_switch_1u_48p", "generic", "switch_1u", "switch", 1),
        ("generic_switch_1u_sfp", "generic", "switch_1u", "switch", 1),
        ("generic_switch_dc_1u", "generic", "switch_dc_1u", "switch", 1),
        ("generic_server_1u", "generic", "server_1u", "server", 1),
        ("generic_server_2u", "generic", "server_2u", "server", 2),
        ("generic_firewall_1u", "generic", "firewall_1u", "firewall", 1),
        ("generic_firewall_2u", "generic", "firewall_2u", "firewall", 2),
        ("generic_ups_2u", "generic", "ups_2u", "ups", 2),
        ("generic_pdu_0u", "generic", "pdu_0u", "pdu", None),
        ("generic_patch_panel_1u", "generic", "patch_panel_1u", "patch_panel", 1),
        ("generic_cable_manager_1u", "generic", "cable_manager_1u", "cable_manager", 1),
        ("generic_blank_1u", "generic", "blank_panel_1u", "blank_panel", 1),
    ]
    for asset_key, maker, family, kind, height in generic_templates:
        grouped.setdefault(asset_key, {
            "asset_key": asset_key,
            "vendor": maker,
            "family": family,
            "exact_models": [],
            "device_type": kind,
            "deployment_domain": "unknown",
            "height_u": height,
            "width_mm": None,
            "depth_mm": None,
            "height_mm": None,
            "dimension_status": "pending_verification",
            "front_orientation": "front",
            "lod": ["LOD0", "LOD1"],
            "glb_path": f"assets/build/glb/{asset_key}.glb",
            "public_path": f"/assets/3d/{asset_key}.glb",
            "thumbnail_path": None,
            "source_blend": f"assets/source/blender/{asset_key}.blend",
            "source_refs": [],
            "data_confidence": "tbd",
            "version": "1.0.0",
            "status": "draft",
            "render_strategy": "procedural_until_glb_approved",
        })

    for record in rack_assets:
        asset_key = record["catalog_key"]
        grouped.setdefault(asset_key, {
            "asset_key": asset_key,
            "vendor": "generic",
            "family": record["asset_model"],
            "exact_models": [record["asset_model"]],
            "device_type": record["category"],
            "deployment_domain": "rack_ecosystem",
            "height_u": None,
            "width_mm": None,
            "depth_mm": None,
            "height_mm": None,
            "dimension_status": "pending_verification",
            "front_orientation": "front",
            "lod": ["LOD0", "LOD1"],
            "glb_path": f"assets/build/glb/{asset_key}.glb",
            "public_path": f"/assets/3d/{asset_key}.glb",
            "thumbnail_path": None,
            "source_blend": f"assets/source/blender/{asset_key}.blend",
            "source_refs": [],
            "data_confidence": record["data_confidence"],
            "version": "1.0.0",
            "status": "draft",
            "render_strategy": "procedural_until_glb_approved",
        })
    return list(grouped.values())


def build_model_mapping(network: list[dict[str, Any]], registry: list[dict[str, Any]]) -> list[dict[str, Any]]:
    registry_keys = {item["asset_key"] for item in registry}
    mappings: list[dict[str, Any]] = []
    for record in network:
        glb_name = clean(record.get("recommended_glb"))
        family_key = slug(Path(glb_name).stem) if glb_name else f"{record['vendor']}__{slug(record['family'])}__generic"
        if family_key not in registry_keys:
            family_key = "generic_switch_1u_24p"
        mappings.append({
            "vendor": record["vendor"],
            "exact_model": record["exact_model"],
            "family": record["family"],
            "asset_key": family_key,
            "resolution_level": "family" if glb_name else "vendor_generic",
            "status": "draft",
            "source_catalog_key": record["catalog_key"],
        })
    return mappings


def yaml_scalar(value: Any) -> str:
    if value is None:
        return "null"
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (int, float)):
        return str(value)
    return json.dumps(str(value), ensure_ascii=False)


def write_yaml(path: Path, records: list[dict[str, Any]], title: str) -> None:
    lines = [
        "# GENERATED FILE - source workbook is preserved separately.",
        f"# {title}",
        "# Review dimension_status/data_confidence before binding a model to production.",
    ]
    for record in records:
        lines.append("-")
        for key, value in record.items():
            lines.append(f"  {key}: {yaml_scalar(value)}")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def load_rows(workbook: Any, index: int) -> list[tuple[int, list[Any]]]:
    sheet = workbook.worksheets[index]
    rows: list[tuple[int, list[Any]]] = []
    # Rows 1-2 are the sheet title and explanation; row 3 is the header.
    for source_row, row in enumerate(sheet.iter_rows(min_row=4, values_only=True), start=4):
        values = list(row)
        if any(clean(value) for value in values):
            rows.append((source_row, values))
    return rows


def build_summary(workbook: Any) -> dict[str, Any]:
    """Keep the workbook's explanation/KPI sheet traceable without editing it."""

    sheet = workbook.worksheets[0]
    rows: list[dict[str, Any]] = []
    for source_row, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        values = list(row)
        if any(clean(value) for value in values):
            rows.append({
                "source_row": source_row,
                "cells": [clean(value) if value is not None else None for value in values],
            })
    return {"sheet_title": sheet.title, "sheet_index": 0, "rows": rows}


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--source", required=True, type=Path)
    parser.add_argument("--output-dir", type=Path, default=Path("assets/catalog"))
    args = parser.parse_args()

    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - only used in offline tooling
        raise SystemExit("openpyxl is required for this offline catalog build") from exc

    source = args.source.resolve()
    if not source.is_file():
        raise SystemExit(f"Workbook not found: {source}")
    workbook = load_workbook(source, data_only=True, read_only=True)
    if len(workbook.worksheets) < 5:
        raise SystemExit("Expected the five RackVision catalog sheets")

    summary = build_summary(workbook)
    network, incomplete_network = build_network(load_rows(workbook, 1))
    rack_assets = build_rack_assets(load_rows(workbook, 2))
    cables = build_cables(load_rows(workbook, 3))
    priorities = build_priorities(load_rows(workbook, 4))
    registry = build_registry(network, rack_assets)
    model_mapping = build_model_mapping(network, registry)

    output_dir = args.output_dir.resolve()
    output_dir.mkdir(parents=True, exist_ok=True)
    write_yaml(output_dir / "network_devices.yaml", network, "Network device catalog")
    write_yaml(output_dir / "rack_assets.yaml", rack_assets, "Rack and accessory catalog")
    write_yaml(output_dir / "cables_optics.yaml", cables, "Cable and optic catalog")
    write_yaml(output_dir / "modeling_priorities.yaml", priorities, "Blender/GLB modeling priority catalog")
    write_yaml(output_dir / "asset_registry.yaml", registry, "Reusable GLB asset registry")
    write_yaml(output_dir / "model_mapping.yaml", model_mapping, "Exact SKU to family/generic asset mapping")
    (output_dir / "catalog_summary.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    # JSON is the machine-readable contract consumed by import/validation
    # tooling.  Keep the per-sheet exports alongside YAML so reviewers can
    # inspect either format without reparsing the source workbook.
    for filename, records in (
        ("network_devices.json", network),
        ("rack_assets.json", rack_assets),
        ("cables_optics.json", cables),
        ("modeling_priorities.json", priorities),
    ):
        (output_dir / filename).write_text(
            json.dumps(records, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )
    (output_dir / "asset_registry.json").write_text(
        json.dumps(registry, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    (output_dir / "model_mapping.json").write_text(
        json.dumps(model_mapping, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )

    digest = hashlib.sha256(source.read_bytes()).hexdigest()
    manifest = {
        "catalog_version": "rackvision-v1",
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "source_file": source.name,
        "source_sha256": digest,
        "source_was_modified": False,
        "summary": {
            "sheet_index": summary["sheet_index"],
            "sheet_title": summary["sheet_title"],
            "records": len(summary["rows"]),
        },
        "sheets": {
            "network_devices": {"sheet_index": 1, "records": len(network), "incomplete_dimensions": incomplete_network},
            "rack_assets": {"sheet_index": 2, "records": len(rack_assets)},
            "cables_optics": {"sheet_index": 3, "records": len(cables)},
            "modeling_priorities": {"sheet_index": 4, "records": len(priorities)},
        },
        "registry": {
            "reusable_asset_entries": len(registry),
            "exact_sku_mappings": len(model_mapping),
            "glb_files_present_at_generation": 0,
        },
        "quality_rules": {
            "unknown_dimensions": "pending_verification",
            "image_based_dimension_inference": False,
            "generic_model_fallback": "allowed only when data_confidence is not verified_official",
        },
    }
    (output_dir / "catalog_manifest.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    (output_dir / "README.txt").write_text(
        "RackVision catalog generated from the supplied workbook.\n"
        "The workbook is not copied or modified by the build utility.\n"
        "catalog_summary.json preserves the explanation/KPI sheet.\n"
        "Review dimension_status and data_confidence before publishing a GLB.\n"
        "Use import_rackvision_device_catalog.py for a dry-run-first device-type import.\n",
        encoding="utf-8",
    )
    print(json.dumps(manifest, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
